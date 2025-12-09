#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Katomo 車両見積書 Excel生成スクリプト
A4印刷対応の高機能見積書システムを作成
"""

import json
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

# データファイルのパス
DATA_DIR = "/Users/ttk/Desktop/katomotor見積もり管理システム/data"
OUTPUT_PATH = "/Users/ttk/Desktop/katomo営業支援ツール/Katomo見積書システム.xlsm"

def load_vehicles():
    """車両データを読み込む"""
    vehicles = []
    with open(os.path.join(DATA_DIR, "vehicles.csv"), 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            vehicles.append(row)
    return vehicles

def load_options():
    """オプションデータを読み込む"""
    options = []
    with open(os.path.join(DATA_DIR, "options.csv"), 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            options.append(row)
    return options

def create_estimate_workbook():
    """見積書Excelファイルを作成"""
    wb = Workbook()

    # シートを作成
    ws_estimate = wb.active
    ws_estimate.title = "見積書"
    ws_data = wb.create_sheet("データ")

    # データを読み込み
    vehicles = load_vehicles()
    options = load_options()

    # データシートを設定
    setup_data_sheet(ws_data, vehicles, options)

    # 見積書シートを設定
    setup_estimate_sheet(ws_estimate, len(vehicles), len(options))

    return wb

def setup_data_sheet(ws, vehicles, options):
    """データシートを設定"""
    # スタイル定義
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 車両マスタテーブル (A1から)
    ws['A1'] = "【車両マスタ】"
    ws['A1'].font = Font(bold=True, size=14)

    vehicle_headers = ['車両名', 'ベース車', '駆動', '排気量', '燃料', '販売価格', '環境性能割', '重量税', '自賠責']
    for col, header in enumerate(vehicle_headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    for row_idx, vehicle in enumerate(vehicles, 3):
        ws.cell(row=row_idx, column=1, value=vehicle.get('modelName', ''))
        ws.cell(row=row_idx, column=2, value=vehicle.get('baseVehicle', ''))
        ws.cell(row=row_idx, column=3, value=vehicle.get('driveType', ''))
        ws.cell(row=row_idx, column=4, value=vehicle.get('engineCc', ''))
        ws.cell(row=row_idx, column=5, value=vehicle.get('fuelType', ''))

        # 価格データ
        price = vehicle.get('sellingPrice', '')
        if price:
            try:
                ws.cell(row=row_idx, column=6, value=float(price))
            except:
                ws.cell(row=row_idx, column=6, value=price)

        tax_env = vehicle.get('taxEnv', '')
        if tax_env:
            try:
                ws.cell(row=row_idx, column=7, value=float(tax_env))
            except:
                ws.cell(row=row_idx, column=7, value=tax_env)

        tax_weight = vehicle.get('taxWeight', '')
        if tax_weight:
            try:
                ws.cell(row=row_idx, column=8, value=float(tax_weight))
            except:
                ws.cell(row=row_idx, column=8, value=tax_weight)

        insurance = vehicle.get('insurance', '')
        if insurance:
            try:
                ws.cell(row=row_idx, column=9, value=float(insurance))
            except:
                ws.cell(row=row_idx, column=9, value=insurance)

        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).border = thin_border

    vehicle_end_row = len(vehicles) + 2

    # オプションマスタテーブル (L1から)
    ws['L1'] = "【オプションマスタ】"
    ws['L1'].font = Font(bold=True, size=14)

    option_headers = ['コード', 'オプション名', 'カテゴリ', '価格']
    for col, header in enumerate(option_headers, 12):  # L列=12
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    for row_idx, option in enumerate(options, 3):
        ws.cell(row=row_idx, column=12, value=option.get('itemCode', ''))
        ws.cell(row=row_idx, column=13, value=option.get('itemName', ''))
        ws.cell(row=row_idx, column=14, value=option.get('category', ''))

        price = option.get('sellingPrice', '')
        if price:
            try:
                ws.cell(row=row_idx, column=15, value=float(price))
            except:
                ws.cell(row=row_idx, column=15, value=price)

        for col in range(12, 16):
            ws.cell(row=row_idx, column=col).border = thin_border

    # 諸費用テーブル (Q1から)
    ws['Q1'] = "【諸費用マスタ】"
    ws['Q1'].font = Font(bold=True, size=14)

    fee_headers = ['費用名', '金額']
    for col, header in enumerate(fee_headers, 17):  # Q列=17
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    fees = [
        ('登録代行手数料', 35000),
        ('車庫証明代行手数料', 15000),
        ('納車費用', 0),
        ('下取り手数料', 15000),
        ('リサイクル料金', 12000),
    ]

    for row_idx, (name, amount) in enumerate(fees, 3):
        ws.cell(row=row_idx, column=17, value=name)
        ws.cell(row=row_idx, column=18, value=amount)
        for col in range(17, 19):
            ws.cell(row=row_idx, column=col).border = thin_border

    # 列幅調整
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['L'].width = 10
    ws.column_dimensions['M'].width = 50
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 14
    ws.column_dimensions['Q'].width = 20
    ws.column_dimensions['R'].width = 14

    # 名前付き範囲を作成（車両リスト用）
    from openpyxl.workbook.defined_name import DefinedName

    # 車両名リストの範囲
    vehicle_range = f"データ!$A$3:$A${vehicle_end_row}"

def setup_estimate_sheet(ws, vehicle_count, option_count):
    """見積書シートを設定"""

    # A4サイズに合わせた列幅設定（全体で約210mm）
    column_widths = {
        'A': 3,    # 左余白
        'B': 5,    # No.
        'C': 35,   # 項目名
        'D': 12,   # 数量
        'E': 14,   # 単価
        'F': 16,   # 金額
        'G': 3,    # 右余白
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # スタイル定義
    title_font = Font(bold=True, size=20)
    subtitle_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    normal_font = Font(size=10)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    thick_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    light_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    # ヘッダー部分
    ws.merge_cells('B2:F2')
    ws['B2'] = "御　見　積　書"
    ws['B2'].font = title_font
    ws['B2'].alignment = Alignment(horizontal='center')

    # 見積日
    ws['E3'] = "見積日:"
    ws['F3'] = ""  # 入力セル
    ws['F3'].fill = yellow_fill
    ws['F3'].border = thin_border

    # 見積番号
    ws['E4'] = "見積No:"
    ws['F4'] = ""  # 入力セル
    ws['F4'].fill = yellow_fill
    ws['F4'].border = thin_border

    # お客様情報
    ws['B6'] = "お客様名:"
    ws.merge_cells('C6:D6')
    ws['C6'] = ""  # 入力セル
    ws['C6'].fill = yellow_fill
    ws['C6'].border = thin_border

    ws['B7'] = "〒"
    ws['C7'] = ""  # 郵便番号入力
    ws['C7'].fill = yellow_fill
    ws['C7'].border = thin_border

    ws['B8'] = "ご住所:"
    ws.merge_cells('C8:F8')
    ws['C8'] = ""  # 住所入力
    ws['C8'].fill = yellow_fill
    ws['C8'].border = thin_border

    ws['B9'] = "TEL:"
    ws['C9'] = ""  # 電話番号入力
    ws['C9'].fill = yellow_fill
    ws['C9'].border = thin_border

    # 会社情報（右側）
    ws['E6'] = "カトモーター"
    ws['E6'].font = Font(bold=True, size=12)

    # 合計金額表示（目立つ）
    ws.merge_cells('B11:C11')
    ws['B11'] = "お見積り金額"
    ws['B11'].font = Font(bold=True, size=14)
    ws['B11'].fill = total_fill
    ws['B11'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('D11:F11')
    ws['D11'] = "=F61"  # 総合計への参照
    ws['D11'].font = Font(bold=True, size=18)
    ws['D11'].fill = total_fill
    ws['D11'].alignment = Alignment(horizontal='right', vertical='center')
    ws['D11'].number_format = '¥#,##0'

    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}11'].border = thick_border

    # テーブルヘッダー
    current_row = 13
    headers = ['No.', '項目', '数量', '単価', '金額']
    cols = ['B', 'C', 'D', 'E', 'F']

    for col, header in zip(cols, headers):
        cell = ws[f'{col}{current_row}']
        cell.value = header
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # 車両情報セクション
    current_row = 14
    ws.merge_cells(f'B{current_row}:F{current_row}')
    ws[f'B{current_row}'] = "【車両本体】"
    ws[f'B{current_row}'].font = Font(bold=True, size=11)
    ws[f'B{current_row}'].fill = light_fill

    # 車両選択行
    current_row = 15
    ws[f'B{current_row}'] = 1
    ws[f'C{current_row}'] = ""  # ドロップダウンで選択
    ws[f'C{current_row}'].fill = yellow_fill
    ws[f'D{current_row}'] = 1
    ws[f'E{current_row}'] = f'=IFERROR(VLOOKUP(C{current_row},データ!$A$3:$F$100,6,FALSE),0)'
    ws[f'F{current_row}'] = f'=D{current_row}*E{current_row}'

    for col in cols:
        ws[f'{col}{current_row}'].border = thin_border
    ws[f'E{current_row}'].number_format = '¥#,##0'
    ws[f'F{current_row}'].number_format = '¥#,##0'

    # 車両選択用ドロップダウン
    dv_vehicle = DataValidation(
        type="list",
        formula1=f'データ!$A$3:$A${vehicle_count + 2}',
        allow_blank=True
    )
    dv_vehicle.error = '車両リストから選択してください'
    dv_vehicle.errorTitle = '無効な入力'
    ws.add_data_validation(dv_vehicle)
    dv_vehicle.add(ws[f'C{current_row}'])

    # オプションセクション
    current_row = 17
    ws.merge_cells(f'B{current_row}:F{current_row}')
    ws[f'B{current_row}'] = "【オプション装備】"
    ws[f'B{current_row}'].font = Font(bold=True, size=11)
    ws[f'B{current_row}'].fill = light_fill

    # オプション行（20行）
    option_start_row = 18
    option_end_row = 37

    # オプション選択用ドロップダウン
    dv_option = DataValidation(
        type="list",
        formula1=f'データ!$M$3:$M${option_count + 2}',
        allow_blank=True
    )
    dv_option.error = 'オプションリストから選択してください'
    dv_option.errorTitle = '無効な入力'
    ws.add_data_validation(dv_option)

    for i, row in enumerate(range(option_start_row, option_end_row + 1), 2):
        ws[f'B{row}'] = i
        ws[f'C{row}'] = ""  # オプション選択
        ws[f'C{row}'].fill = yellow_fill
        ws[f'D{row}'] = ""  # 数量入力
        ws[f'D{row}'].fill = yellow_fill
        ws[f'E{row}'] = f'=IFERROR(VLOOKUP(C{row},データ!$M$3:$O$200,3,FALSE),0)'
        ws[f'F{row}'] = f'=IF(D{row}="",0,D{row}*E{row})'

        for col in cols:
            ws[f'{col}{row}'].border = thin_border
        ws[f'E{row}'].number_format = '¥#,##0'
        ws[f'F{row}'].number_format = '¥#,##0'

        dv_option.add(ws[f'C{row}'])

    # オプション小計
    current_row = 38
    ws.merge_cells(f'B{current_row}:D{current_row}')
    ws[f'B{current_row}'] = "オプション小計"
    ws[f'B{current_row}'].alignment = Alignment(horizontal='right')
    ws[f'B{current_row}'].font = header_font
    ws[f'F{current_row}'] = f'=SUM(F{option_start_row}:F{option_end_row})'
    ws[f'F{current_row}'].number_format = '¥#,##0'
    ws[f'F{current_row}'].font = header_font

    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{current_row}'].border = thin_border

    # 諸費用セクション
    current_row = 40
    ws.merge_cells(f'B{current_row}:F{current_row}')
    ws[f'B{current_row}'] = "【税金・諸費用】"
    ws[f'B{current_row}'].font = Font(bold=True, size=11)
    ws[f'B{current_row}'].fill = light_fill

    # 税金・諸費用項目
    fees = [
        ('環境性能割', f'=IFERROR(VLOOKUP(C15,データ!$A$3:$G$100,7,FALSE),0)'),
        ('重量税', f'=IFERROR(VLOOKUP(C15,データ!$A$3:$H$100,8,FALSE),0)'),
        ('自賠責保険', f'=IFERROR(VLOOKUP(C15,データ!$A$3:$I$100,9,FALSE),0)'),
        ('登録代行手数料', '35000'),
        ('車庫証明代行手数料', '15000'),
        ('納車費用', ''),
        ('下取り手数料', ''),
        ('リサイクル料金', '12000'),
    ]

    fee_start_row = 41
    for i, (name, formula) in enumerate(fees):
        row = fee_start_row + i
        ws[f'C{row}'] = name
        if formula.startswith('='):
            ws[f'F{row}'] = formula
        else:
            ws[f'F{row}'] = float(formula) if formula else 0
            ws[f'F{row}'].fill = yellow_fill  # 編集可能
        ws[f'F{row}'].number_format = '¥#,##0'
        ws[f'F{row}'].border = thin_border

    fee_end_row = fee_start_row + len(fees) - 1

    # 諸費用小計
    current_row = fee_end_row + 1
    ws.merge_cells(f'B{current_row}:D{current_row}')
    ws[f'B{current_row}'] = "諸費用小計"
    ws[f'B{current_row}'].alignment = Alignment(horizontal='right')
    ws[f'B{current_row}'].font = header_font
    ws[f'F{current_row}'] = f'=SUM(F{fee_start_row}:F{fee_end_row})'
    ws[f'F{current_row}'].number_format = '¥#,##0'
    ws[f'F{current_row}'].font = header_font

    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{current_row}'].border = thin_border

    # 下取り・値引きセクション
    current_row = 52
    ws.merge_cells(f'B{current_row}:F{current_row}')
    ws[f'B{current_row}'] = "【下取り・値引き】"
    ws[f'B{current_row}'].font = Font(bold=True, size=11)
    ws[f'B{current_row}'].fill = light_fill

    adjustments = [
        ('下取り車', ''),
        ('下取り価格', ''),
        ('値引き', ''),
    ]

    adj_start_row = 53
    for i, (name, default) in enumerate(adjustments):
        row = adj_start_row + i
        ws[f'C{row}'] = name
        ws[f'F{row}'] = default if default else 0
        ws[f'F{row}'].fill = yellow_fill
        ws[f'F{row}'].number_format = '¥#,##0'
        ws[f'F{row}'].border = thin_border

    # 消費税セクション
    current_row = 57
    ws.merge_cells(f'B{current_row}:D{current_row}')
    ws[f'B{current_row}'] = "車両本体・オプション合計（税抜）"
    ws[f'B{current_row}'].alignment = Alignment(horizontal='right')
    ws[f'F{current_row}'] = f'=F15+F38'
    ws[f'F{current_row}'].number_format = '¥#,##0'

    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{current_row}'].border = thin_border

    current_row = 58
    ws.merge_cells(f'B{current_row}:D{current_row}')
    ws[f'B{current_row}'] = "消費税（10%）"
    ws[f'B{current_row}'].alignment = Alignment(horizontal='right')
    ws[f'F{current_row}'] = f'=ROUND(F57*0.1,0)'
    ws[f'F{current_row}'].number_format = '¥#,##0'

    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{current_row}'].border = thin_border

    # 総合計
    current_row = 61
    ws.merge_cells(f'B{current_row}:D{current_row}')
    ws[f'B{current_row}'] = "総合計"
    ws[f'B{current_row}'].font = Font(bold=True, size=14)
    ws[f'B{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'B{current_row}'].fill = total_fill

    ws[f'F{current_row}'] = f'=F57+F58+F49-F54-F55+F53'
    ws[f'F{current_row}'].font = Font(bold=True, size=16)
    ws[f'F{current_row}'].number_format = '¥#,##0'
    ws[f'F{current_row}'].fill = total_fill

    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{current_row}'].border = thick_border

    # 備考欄
    current_row = 63
    ws[f'B{current_row}'] = "備考:"
    ws[f'B{current_row}'].font = header_font

    ws.merge_cells(f'B64:F68')
    ws['B64'] = ""
    ws['B64'].fill = yellow_fill
    ws['B64'].border = thin_border
    ws['B64'].alignment = Alignment(wrap_text=True, vertical='top')

    # フッター
    current_row = 70
    ws[f'B{current_row}'] = "※本見積書の有効期限は発行日より1ヶ月間です。"
    ws[f'B{current_row}'].font = Font(size=9)

    current_row = 71
    ws[f'B{current_row}'] = "※価格は予告なく変更される場合があります。"
    ws[f'B{current_row}'].font = Font(size=9)

    # 印刷設定
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_area = 'A1:G72'

    # 印刷時の余白設定
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

def main():
    print("Katomo見積書システムを作成中...")

    # scriptsディレクトリを作成
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    wb = create_estimate_workbook()

    # .xlsm形式で保存（VBAマクロ用）
    # 注意: openpyxlではVBAマクロを直接追加できないため、
    # まず.xlsxで保存し、後でExcelでVBAを追加する必要があります
    xlsx_path = OUTPUT_PATH.replace('.xlsm', '.xlsx')
    wb.save(xlsx_path)

    print(f"Excelファイルを保存しました: {xlsx_path}")
    print("\n【次のステップ】")
    print("1. Excelでファイルを開く")
    print("2. Alt+F11でVBAエディタを開く")
    print("3. 以下のVBAマクロを追加:")
    print("""
    Sub ClearEstimate()
        ' 見積もりクリアマクロ
        With Worksheets("見積書")
            .Range("C6").ClearContents  ' お客様名
            .Range("C7").ClearContents  ' 郵便番号
            .Range("C8").ClearContents  ' 住所
            .Range("C9").ClearContents  ' 電話番号
            .Range("F3").ClearContents  ' 見積日
            .Range("F4").ClearContents  ' 見積番号
            .Range("C15").ClearContents ' 車両選択
            .Range("C18:C37").ClearContents ' オプション
            .Range("D18:D37").ClearContents ' 数量
            .Range("F44:F48").ClearContents ' 諸費用（編集可能部分）
            .Range("F53:F55").ClearContents ' 下取り・値引き
            .Range("B64").ClearContents ' 備考
        End With
        MsgBox "見積書をクリアしました", vbInformation
    End Sub

    Sub ExportToPDF()
        ' PDF出力マクロ
        Dim pdfPath As String
        Dim customerName As String
        Dim estimateNo As String

        customerName = Worksheets("見積書").Range("C6").Value
        estimateNo = Worksheets("見積書").Range("F4").Value

        If customerName = "" Then customerName = "お客様"
        If estimateNo = "" Then estimateNo = Format(Now, "yyyymmdd_hhnnss")

        pdfPath = Application.GetSaveAsFilename( _
            InitialFileName:=customerName & "_見積書_" & estimateNo & ".pdf", _
            FileFilter:="PDFファイル (*.pdf), *.pdf")

        If pdfPath <> "False" Then
            Worksheets("見積書").ExportAsFixedFormat _
                Type:=xlTypePDF, _
                Filename:=pdfPath, _
                Quality:=xlQualityStandard, _
                IncludeDocProperties:=True, _
                IgnorePrintAreas:=False
            MsgBox "PDFを保存しました:" & vbCrLf & pdfPath, vbInformation
        End If
    End Sub
    """)
    print("4. .xlsm形式で保存")

if __name__ == "__main__":
    main()
